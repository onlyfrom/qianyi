from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, send_file, send_from_directory
from decimal import Decimal
from .model import PurchaseOrder, PurchaseOrderItem, User, db, DeliveryOrder, DeliveryItem, Payment, Product, UserProductPrice
from .views import login_required, app
import pandas as pd
from io import BytesIO
from PIL import Image, ImageDraw, ImageFont
import os
import time
import logging
import requests
import traceback
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)
WX_ENV = 'prod-9gd4jllic76d4842'

# 恢复使用蓝图，确保登录状态正确继承
billing_bp = Blueprint('billing', __name__)

@billing_bp.route('/billing', methods=['GET'])
@login_required
def get_billing_list(user_id):
    """获取账单列表 - 按客户汇总"""
    try:
        # 从URL参数获取数据
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        customer_id = request.args.get('customer_id')
        sort_by = request.args.get('sort_by', 'total_amount')  # 默认按货款总额排序
        sort_order = request.args.get('sort_order', 'desc')    # 默认降序

        print(f"[DEBUG] 获取账单列表: 页码={page}, 每页数量={page_size}, 开始日期={start_date}, 结束日期={end_date}, 客户ID={customer_id}, 排序字段={sort_by}, 排序方式={sort_order}")

        # 构建基础查询 - 按客户分组汇总
        base_query = db.session.query(
            User.id.label('customer_id'),
            User.nickname.label('customer_name'),
            db.func.sum(DeliveryItem.quantity).label('total_quantity'),  # 使用发货单数量
            db.func.sum(PurchaseOrderItem.price * DeliveryItem.quantity).label('total_amount'),  # 使用发货单数量计算金额
            db.func.sum(DeliveryOrder.additional_fee).label('total_additional_fee'),  # 添加附加费用汇总
            db.func.sum(PurchaseOrder.paid_amount).label('paid_amount')
        ).join(
            PurchaseOrder, User.id == PurchaseOrder.user_id
        ).join(
            PurchaseOrderItem, PurchaseOrder.id == PurchaseOrderItem.order_id
        ).join(
            DeliveryItem, 
            db.and_(
                DeliveryItem.order_number == PurchaseOrder.order_number,
                DeliveryItem.product_id == PurchaseOrderItem.product_id,
                DeliveryItem.color == PurchaseOrderItem.color
            )
        ).join(
            DeliveryOrder,
            db.and_(
                DeliveryOrder.id == DeliveryItem.delivery_id,
                DeliveryOrder.status.in_([1, 2])  # 只统计已发货和已完成的订单
            )
        ).group_by(
            User.id, User.nickname
        )

        # 应用筛选条件
        if start_date:
            try:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                base_query = base_query.filter(DeliveryOrder.created_at >= start_datetime)  # 使用发货单日期
                print(f"[DEBUG] 应用开始日期筛选: {start_datetime}")
            except ValueError as e:
                print(f"[ERROR] 开始日期格式错误: {start_date}, 错误: {str(e)}")
                return jsonify({'error': '开始日期格式错误，请使用YYYY-MM-DD格式'}), 400
                
        if end_date:
            try:
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                end_datetime = end_datetime + timedelta(days=1) - timedelta(seconds=1)
                base_query = base_query.filter(DeliveryOrder.created_at <= end_datetime)  # 使用发货单日期
                print(f"[DEBUG] 应用结束日期筛选: {end_datetime}")
            except ValueError as e:
                print(f"[ERROR] 结束日期格式错误: {end_date}, 错误: {str(e)}")
                return jsonify({'error': '结束日期格式错误，请使用YYYY-MM-DD格式'}), 400
                
        if customer_id:
            base_query = base_query.filter(User.id == customer_id)

        # 获取总数
        total = base_query.count()
        
        # 排序
        sort_column = None
        if sort_by == 'customer_name':
            sort_column = User.nickname
        elif sort_by == 'total_quantity':
            sort_column = db.func.sum(DeliveryItem.quantity)
        elif sort_by == 'total_amount':
            sort_column = db.func.sum(PurchaseOrderItem.price * DeliveryItem.quantity)
        elif sort_by == 'paid_amount':
            sort_column = db.func.sum(PurchaseOrder.paid_amount)
        elif sort_by == 'unpaid_amount':
            sort_column = db.func.sum(PurchaseOrderItem.price * PurchaseOrderItem.quantity) - db.func.sum(PurchaseOrder.paid_amount)
        
        # 应用排序
        if sort_column is not None:
            if sort_order == 'desc':
                base_query = base_query.order_by(db.desc(sort_column))
            else:
                base_query = base_query.order_by(sort_column)
        else:
            # 默认排序
            base_query = base_query.order_by(db.desc(db.func.sum(PurchaseOrderItem.price * PurchaseOrderItem.quantity)))
        
        # 分页
        base_query = base_query.offset((page - 1) * page_size).limit(page_size)

        # 执行查询
        results = base_query.all()

        # 构建返回数据
        items = []
        for result in results:
            customer_id, customer_name, total_quantity, total_amount, total_additional_fee, paid_amount = result
            # 重新查询该客户的所有发货单附加费用总和
            additional_fee_query = db.session.query(
                db.func.sum(DeliveryOrder.additional_fee)
            ).filter(
                DeliveryOrder.customer_id == customer_id,
                DeliveryOrder.status.in_([1, 2])  # 只统计已发货和已完成的订单
            ).scalar()

            total_additional_fee = float(additional_fee_query or 0)
            actual_total = float(total_amount or 0) + total_additional_fee

            items.append({
                'id': customer_id,  # 使用客户ID作为标识
                'customer_name': customer_name,
                'total_quantity': total_quantity or 0,
                'total_amount': actual_total,  # 包含附加费用的总金额
                'additional_fee': total_additional_fee,  # 添加附加费用字段
                'paid_amount': float(paid_amount or 0),
                'unpaid_amount': float(actual_total - (paid_amount or 0))  # 使用包含附加费用的总金额计算未付金额
            })

        print(f"[DEBUG] 成功获取账单列表: 总数={total}, 当前页数量={len(items)}")
        return jsonify({
            'items': items,
            'total': total
        })
    except Exception as e:
        print(f"[ERROR] 获取账单列表失败: {str(e)}")
        return jsonify({'error': '获取账单列表失败'}), 500

@billing_bp.route('/billing/statistics', methods=['GET'])
@login_required
def get_billing_statistics(user_id):
    """获取账单统计数据"""
    try:
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 构建订单总额和数量查询 - 只统计已发货的订单
        order_query = db.session.query(
            db.func.sum(DeliveryItem.quantity).label('total_count'),
            db.func.sum(
                db.case(
                    (DeliveryOrder.status.in_([1, 2]), PurchaseOrderItem.price * DeliveryItem.quantity),
                    else_=0
                )
            ).label('total_amount'),
            db.func.sum(
                db.case(
                    (DeliveryOrder.status.in_([1, 2]), DeliveryOrder.additional_fee),
                    else_=0
                )
            ).label('total_additional_fee')
        ).join(
            DeliveryOrder, DeliveryItem.delivery_id == DeliveryOrder.id
        ).join(
            PurchaseOrder, DeliveryItem.order_number == PurchaseOrder.order_number
        ).join(
            PurchaseOrderItem,
            db.and_(
                PurchaseOrder.id == PurchaseOrderItem.order_id,
                PurchaseOrderItem.product_id == DeliveryItem.product_id,
                PurchaseOrderItem.color == DeliveryItem.color
            )
        )
        
        # 构建收款金额查询
        payment_query = db.session.query(
            db.func.sum(Payment.amount).label('paid_amount')
        )
        
        # 应用日期筛选
        if start_date:
            try:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                order_query = order_query.filter(DeliveryOrder.created_at >= start_datetime)
                payment_query = payment_query.filter(Payment.payment_date >= start_datetime)
            except ValueError:
                return jsonify({'error': '开始日期格式错误，请使用YYYY-MM-DD格式'}), 400
                
        if end_date:
            try:
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                end_datetime = end_datetime + timedelta(days=1) - timedelta(seconds=1)
                order_query = order_query.filter(DeliveryOrder.created_at <= end_datetime)
                payment_query = payment_query.filter(Payment.payment_date <= end_datetime)
            except ValueError:
                return jsonify({'error': '结束日期格式错误，请使用YYYY-MM-DD格式'}), 400
        
        # 执行查询
        order_result = order_query.first()
        payment_result = payment_query.scalar() or 0
        
        # 计算统计数据
        total_amount = float(order_result.total_amount or 0)
        total_additional_fee = float(order_result.total_additional_fee or 0)
        actual_total = total_amount + total_additional_fee
        paid_amount = float(payment_result)
        unpaid_amount = actual_total - paid_amount
        
        return jsonify({
            'totalAmount': actual_total,
            'paidAmount': paid_amount,
            'unpaidAmount': unpaid_amount,
            'totalCount': order_result.total_count or 0,
            'totalAdditionalFee': total_additional_fee
        })
    except Exception as e:
        print(f"[ERROR] 获取账单统计数据失败: {str(e)}")
        return jsonify({'error': '获取账单统计数据失败'}), 500

@billing_bp.route('/billing/<int:customer_id>/orders', methods=['GET'])
@login_required
def get_customer_orders(user_id, customer_id):
    """获取客户的发货单列表"""
    try:
        print(f"[DEBUG] 获取客户发货单列表: 客户ID={customer_id}")
        
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        keyword = request.args.get('keyword')
        
        # 获取客户的已收款总额
        total_paid = db.session.query(
            db.func.sum(Payment.amount)
        ).filter(
            Payment.customer_id == customer_id
        ).scalar() or 0
        
        # 构建基础查询
        query = db.session.query(
            DeliveryOrder,
            db.func.sum(DeliveryItem.quantity).label('total_quantity'),
            db.func.sum(PurchaseOrderItem.price * DeliveryItem.quantity).label('total_amount'),
            DeliveryOrder.additional_fee.label('additional_fee')  # 添加附加费用字段
        ).join(
            DeliveryItem, DeliveryOrder.id == DeliveryItem.delivery_id
        ).join(
            PurchaseOrder, DeliveryItem.order_number == PurchaseOrder.order_number
        ).join(
            PurchaseOrderItem, 
            db.and_(
                PurchaseOrder.id == PurchaseOrderItem.order_id,
                PurchaseOrderItem.product_id == DeliveryItem.product_id,
                PurchaseOrderItem.color == DeliveryItem.color
            )
        ).join(
            Product, Product.id == PurchaseOrderItem.product_id
        ).filter(
            DeliveryOrder.customer_id == customer_id
        )
        
        # 添加日期过滤
        if start_date:
            query = query.filter(db.func.date(DeliveryOrder.created_at) >= start_date)
        if end_date:
            query = query.filter(db.func.date(DeliveryOrder.created_at) <= end_date)
            
        # 添加关键词搜索
        if keyword:
            query = query.filter(
                db.or_(
                    DeliveryOrder.order_number.ilike(f'%{keyword}%'),
                    Product.name.ilike(f'%{keyword}%'),
                    DeliveryOrder.remark.ilike(f'%{keyword}%')
                )
            )
            
        # 分组并按创建时间降序排序
        query = query.group_by(DeliveryOrder.id, DeliveryOrder.created_at).order_by(DeliveryOrder.created_at.desc())
        
        # 执行查询
        orders = query.all()
        
        # 计算总发货金额（包含附加费用）
        total_delivery_amount = sum(float(order[2] or 0) for order in orders)
        total_additional_fee = sum(float(order[3] or 0) for order in orders)
        total_amount = total_delivery_amount + total_additional_fee

        # 计算商品总数
        total_all_quantity = sum(float(order[1] or 0) for order in orders)
        
        # 构建返回数据
        result = []
        total_amount = 0  # 初始化总金额
        for order, total_quantity, order_amount, additional_fee in orders:
            # 获取创建者信息
            creator = User.query.get(order.created_by)
            
            # 状态文本映射
            status_text_map = {
                0: '已开单',
                1: '已发货',
                2: '已完成',
                3: '已取消',
                4: '异常'
            }
            
            # 计算订单总金额（包含附加费用）
            order_total = float(order_amount or 0) + float(additional_fee or 0)
            total_amount += order_total  # 累加每个订单的总金额
            
            result.append({
                'id': order.id,
                'order_number': order.order_number,
                'created_at': order.created_at.strftime('%Y-%m-%d'),
                'total_quantity': total_quantity or 0,
                'total_amount': order_total,
                'additional_fee': float(additional_fee or 0),
                'status': order.status,
                'status_text': status_text_map.get(order.status, '未知状态'),
                'logistics_company': order.logistics_company,
                'tracking_number': order.tracking_number,
                'remark': order.remark,
                'creator': {
                    'id': creator.id if creator else None,
                    'username': creator.username if creator else None,
                    'nickname': creator.nickname if creator else None
                } if creator else None
            })
            
        return jsonify({
            'code': 0,
            'data': result,
            'total_paid': total_paid,
            'total_amount': total_amount,
            'total_additional_fee': total_additional_fee,
            'total_quantity': total_all_quantity
        })
        
    except Exception as e:
        print(f"[ERROR] 获取客户发货单列表失败: 客户ID={customer_id}, 错误={str(e)}")
        return jsonify({
            'code': -1,
            'message': f'获取客户发货单列表失败: {str(e)}'
        })

@billing_bp.route('/billing/<int:order_id>/payment', methods=['POST'])
@login_required
def confirm_payment(user_id, order_id):
    """确认收款"""
    try:
        print(f"[DEBUG] 确认收款: 订单ID={order_id}")
        data = request.get_json()
        amount = float(data.get('amount', 0))
        remark = data.get('remark', '')

        if amount <= 0:
            return jsonify({'error': '收款金额必须大于0'}), 400

        order = PurchaseOrder.query.get_or_404(order_id)
        
        # 计算订单总金额
        total_amount = db.session.query(
            db.func.sum(PurchaseOrderItem.price * PurchaseOrderItem.quantity)
        ).filter_by(order_id=order_id).scalar() or 0

        # 检查收款金额是否超过未付金额
        unpaid_amount = total_amount - (order.paid_amount or 0)
        if amount > unpaid_amount:
            return jsonify({'error': '收款金额不能超过未付金额'}), 400

        # 更新已付金额
        order.paid_amount = (order.paid_amount or 0) + amount
        if remark:
            order.remark = remark

        # 更新订单状态
        if order.paid_amount >= total_amount:
            order.status = 2  # 设置为已结清
        elif order.paid_amount > 0:
            order.status = 1  # 设置为部分付款
        else:
            order.status = 0  # 设置为未付款

        db.session.commit()
        
        print(f"[DEBUG] 成功确认收款: 订单ID={order_id}, 收款金额={amount}")
        return jsonify({'message': '收款确认成功'})
    except Exception as e:
        print(f"[ERROR] 确认收款失败: 订单ID={order_id}, 错误={str(e)}")
        db.session.rollback()
        return jsonify({'error': '确认收款失败'}), 500

@billing_bp.route('/billing/export', methods=['GET'])
@login_required
def export_billing(user_id):
    """导出账单数据"""
    try:
        print("[DEBUG] 开始导出账单数据")
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        customer_id = request.args.get('customer_id')

        # 构建查询条件 - 按客户分组汇总
        base_query = db.session.query(
            User.id.label('customer_id'),
            User.nickname.label('customer_name'),
            db.func.sum(DeliveryItem.quantity).label('total_quantity'),  # 使用发货单数量
            db.func.sum(PurchaseOrderItem.price * DeliveryItem.quantity).label('total_amount'),  # 使用发货单数量计算金额
            db.func.sum(DeliveryOrder.additional_fee).label('total_additional_fee'),  # 添加附加费用汇总
            db.func.sum(PurchaseOrder.paid_amount).label('paid_amount')
        ).join(
            PurchaseOrder, User.id == PurchaseOrder.user_id
        ).join(
            PurchaseOrderItem, PurchaseOrder.id == PurchaseOrderItem.order_id
        ).join(
            DeliveryItem, 
            db.and_(
                DeliveryItem.order_number == PurchaseOrder.order_number,
                DeliveryItem.product_id == PurchaseOrderItem.product_id,
                DeliveryItem.color == PurchaseOrderItem.color
            )
        ).join(
            DeliveryOrder,
            db.and_(
                DeliveryOrder.id == DeliveryItem.delivery_id,
                DeliveryOrder.status.in_([1, 2])  # 只统计已发货和已完成的订单
            )
        ).group_by(
            User.id, User.nickname
        )

        # 添加日期过滤
        if start_date:
            try:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                base_query = base_query.filter(DeliveryOrder.created_at >= start_datetime)  # 使用发货单日期
            except ValueError:
                return jsonify({'error': '开始日期格式错误，请使用YYYY-MM-DD格式'}), 400
                
        if end_date:
            try:
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                end_datetime = end_datetime + timedelta(days=1) - timedelta(seconds=1)
                base_query = base_query.filter(DeliveryOrder.created_at <= end_datetime)  # 使用发货单日期
            except ValueError:
                return jsonify({'error': '结束日期格式错误，请使用YYYY-MM-DD格式'}), 400

        # 添加客户过滤
        if customer_id:
            base_query = base_query.filter(User.id == customer_id)

        # 执行查询
        results = base_query.all()

        # 构建Excel数据
        data = []
        for result in results:
            customer_id, customer_name, total_quantity, total_amount, total_additional_fee, paid_amount = result
            # 重新查询该客户的所有发货单附加费用总和
            additional_fee_query = db.session.query(
                db.func.sum(DeliveryOrder.additional_fee)
            ).filter(
                DeliveryOrder.customer_id == customer_id,
                DeliveryOrder.status.in_([1, 2])
            ).scalar()

            total_additional_fee = float(additional_fee_query or 0)
            actual_total = float(total_amount or 0) + total_additional_fee

            data.append({
                '客户ID': customer_id,
                '客户名称': customer_name,
                '商品总数': total_quantity or 0,
                '商品总额': float(total_amount or 0),
                '附加费用': total_additional_fee,
                '应付总额': actual_total,
                '已付金额': float(paid_amount or 0),
                '未付金额': float(actual_total - (paid_amount or 0))
            })

        df = pd.DataFrame(data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='账单数据')

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'账单数据_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        print(f"[ERROR] 导出账单数据失败: {str(e)}")
        return jsonify({'error': '导出账单数据失败'}), 500

@billing_bp.route('/billing/<int:customer_id>/delivery_orders', methods=['GET'])
@login_required
def get_customer_delivery_orders(user_id, customer_id):
    """获取客户的发货单列表"""
    try:
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 构建基础查询
        base_query = db.session.query(
            DeliveryOrder.id,
            DeliveryOrder.order_number,
            DeliveryOrder.created_at,
            DeliveryOrder.status,
            DeliveryOrder.remark,
            db.func.sum(DeliveryItem.quantity).label('total_quantity')
        ).join(
            DeliveryItem, DeliveryOrder.id == DeliveryItem.delivery_id
        ).filter(
            DeliveryOrder.customer_id == customer_id
        ).group_by(
            DeliveryOrder.id,
            DeliveryOrder.order_number,
            DeliveryOrder.created_at,
            DeliveryOrder.status,
            DeliveryOrder.remark
        )
        
        # 应用日期筛选
        if start_date:
            try:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                base_query = base_query.filter(DeliveryOrder.created_at >= start_datetime)
            except ValueError:
                return jsonify({'error': '开始日期格式错误，请使用YYYY-MM-DD格式'}), 400
                
        if end_date:
            try:
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                end_datetime = end_datetime + timedelta(days=1) - timedelta(seconds=1)
                base_query = base_query.filter(DeliveryOrder.created_at <= end_datetime)
            except ValueError:
                return jsonify({'error': '结束日期格式错误，请使用YYYY-MM-DD格式'}), 400
        
        # 执行查询
        results = base_query.all()
        
        # 构建返回数据
        delivery_orders = []
        for result in results:
            order_id, order_number, created_at, status, remark, total_quantity = result
            
            # 获取创建者信息
            creator = User.query.get(DeliveryOrder.query.get(order_id).created_by)
            
            # 状态文本映射
            status_text_map = {
                0: '已开单',
                1: '已发货',
                2: '已完成',
                3: '已取消',
                4: '异常'
            }
            
            delivery_orders.append({
                'id': order_id,
                'order_number': order_number,
                'created_at': created_at.isoformat(),
                'status': status,
                'status_text': status_text_map.get(status, '未知状态'),
                'remark': remark,
                'total_quantity': total_quantity or 0,
                'creator': {
                    'id': creator.id,
                    'username': creator.username,
                    'nickname': creator.nickname
                } if creator else None
            })
        
        return jsonify({
            'code': 0,
            'data': delivery_orders
        })
        
    except Exception as e:
        print(f"[ERROR] 获取客户发货单列表失败: {str(e)}")
        return jsonify({'error': '获取客户发货单列表失败'}), 500

@billing_bp.route('/payments', methods=['POST'])
@login_required
def create_payment(user_id):
    """创建收款记录"""
    try:
        data = request.get_json()
        customer_id = data.get('customer_id')
        customer_name = data.get('customer_name')
        amount = float(data.get('amount', 0))
        payment_date = datetime.strptime(data.get('payment_date'), '%Y-%m-%d %H:%M:%S')
        remark = data.get('remark', '')
        delivery_orders = data.get('delivery_orders', [])

        print(f"[DEBUG] 创建收款记录: 客户ID={customer_id}, 金额={amount}, 时间={payment_date}")

        if amount <= 0:
            return jsonify({'error': '收款金额必须大于0'}), 400

        # 创建收款记录
        payment = Payment(
            customer_id=customer_id,
            customer_name=customer_name,
            amount=amount,
            payment_date=payment_date,
            remark=remark,
            delivery_orders=delivery_orders
        )
        db.session.add(payment)

        # 获取客户所有未结清的订单
        orders = PurchaseOrder.query.filter(
            PurchaseOrder.user_id == customer_id,
            PurchaseOrder.status.in_([0, 1])  # 未付款或部分付款的订单
        ).all()

        remaining_amount = amount
        for order in orders:
            if remaining_amount <= 0:
                break

            # 计算订单总金额
            total_amount = db.session.query(
                db.func.sum(PurchaseOrderItem.price * PurchaseOrderItem.quantity)
            ).filter_by(order_id=order.id).scalar() or 0

            # 计算订单未付金额
            unpaid_amount = total_amount - (order.paid_amount or 0)
            
            if unpaid_amount > 0:
                # 确定本次为该订单付款的金额
                payment_for_order = min(remaining_amount, unpaid_amount)
                
                # 更新订单已付金额
                order.paid_amount = (order.paid_amount or 0) + payment_for_order
                
                # 更新订单状态
                if order.paid_amount >= total_amount:
                    order.status = 2  # 已结清
                else:
                    order.status = 1  # 部分付款
                
                remaining_amount -= payment_for_order

        # 如果指定了发货单，更新发货单状态
        if delivery_orders:
            for delivery_id in delivery_orders:
                delivery_order = DeliveryOrder.query.get(delivery_id)
                if delivery_order:
                    delivery_order.status = 2  # 设置为已完成状态

        db.session.commit()
        
        print(f"[DEBUG] 成功创建收款记录: 客户ID={customer_id}, 金额={amount}")
        return jsonify({'message': '收款记录创建成功'}), 201

    except Exception as e:
        print(f"[ERROR] 创建收款记录失败: {str(e)}")
        db.session.rollback()
        return jsonify({'error': f'创建收款记录失败: {str(e)}'}), 500

@billing_bp.route('/payments/<int:customer_id>/history', methods=['GET'])
@login_required
def get_payment_history(user_id, customer_id):
    """获取客户的收款历史记录"""
    try:
        print(f"[DEBUG] 获取收款历史: 客户ID={customer_id}")
        
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 构建基础查询
        base_query = Payment.query.filter_by(customer_id=customer_id)
        
        # 添加日期过滤
        if start_date:
            base_query = base_query.filter(Payment.payment_date >= start_date)
        if end_date:
            base_query = base_query.filter(Payment.payment_date <= end_date)
            
        # 按支付时间倒序排序
        base_query = base_query.order_by(Payment.payment_date.desc())
        
        # 执行查询
        payments = base_query.all()
        
        # 构建返回数据
        payment_history = [payment.to_dict() for payment in payments]
            
        return jsonify({
            'code': 0,
            'data': payment_history
        })
        
    except Exception as e:
        print(f"[ERROR] 获取收款历史失败: {str(e)}")
        return jsonify({
            'code': -1,
            'message': f'获取收款历史失败: {str(e)}'
        })

@billing_bp.route('/billing/<int:delivery_id>/image', methods=['GET'])
@login_required
def generate_delivery_image(user_id, delivery_id):
    """生成发货单图片"""
    try:
        # 获取发货单信息
        delivery_order = DeliveryOrder.query.get_or_404(delivery_id)
        logger.info(f"开始生成发货单图片: delivery_id={delivery_id}, order_number={delivery_order.order_number}")
        
        # 获取发货单商品
        delivery_items = DeliveryItem.query.filter_by(delivery_id=delivery_id).all()
        if not delivery_items:
            logger.error(f"发货单商品列表为空: delivery_id={delivery_id}")
            return jsonify({'error': '发货单商品列表为空'}), 400
        
        # 获取对应的采购单
        purchase_order = PurchaseOrder.query.filter_by(order_number=delivery_order.order_number).first()
        if not purchase_order:
            logger.error(f"未找到对应的采购单: order_number={delivery_order.order_number}")
            return jsonify({'error': '未找到对应的采购单'}), 404
            
        # 获取客户的所有发货单总额
        total_delivery_amount = db.session.query(
            db.func.sum(PurchaseOrderItem.price * DeliveryItem.quantity)
        ).join(
            PurchaseOrder,
            PurchaseOrder.id == PurchaseOrderItem.order_id
        ).join(
            DeliveryItem, 
            db.and_(
                DeliveryItem.order_number == PurchaseOrder.order_number,
                DeliveryItem.product_id == PurchaseOrderItem.product_id,
                DeliveryItem.color == PurchaseOrderItem.color
            )
        ).join(
            DeliveryOrder,
            DeliveryOrder.id == DeliveryItem.delivery_id
        ).join(
            User,
            User.id == PurchaseOrder.user_id
        ).filter(
            User.id == purchase_order.user_id
        ).scalar() or 0

        # 获取客户的已收款总额
        total_paid = db.session.query(
            db.func.sum(Payment.amount)
        ).join(
            User,
            User.id == Payment.customer_id
        ).filter(
            User.id == purchase_order.user_id
        ).scalar() or 0

        # 计算累计应付总额
        total_unpaid = float(total_delivery_amount) - float(total_paid)
        logger.info(f"客户累计应付总额: customer_id={purchase_order.user_id}, total_delivery={total_delivery_amount}, total_paid={total_paid}, total_unpaid={total_unpaid}")

        # 计算总数量和总金额
        total_quantity = sum(item.quantity for item in delivery_items)
        total_amount = 0
        
        # 创建商品单价映射
        price_map = {}
        for po_item in purchase_order.items:
            key = (po_item.product_id, po_item.color)
            price_map[key] = {
                'price': po_item.price,
                'logo_price': po_item.logo_price,
                'accessory_price': po_item.accessory_price,
                'packaging_price': po_item.packaging_price
            }

        # 计算不同的包裹数
        package_count = len(set(item.package_id for item in delivery_items if item.package_id))
        logger.info(f"计算包裹数: package_count={package_count}")

        try:
            # 创建图片
            image = Image.new('RGB', (800, 1200), 'white')
            draw = ImageDraw.Draw(image)
            
            # 加载字体
            try:
                font_path = os.path.join(os.path.dirname(__file__), 'static', 'simsun.ttc')
                font = ImageFont.truetype(font_path, 24)
                title_font = ImageFont.truetype(font_path, 32)
                small_font = ImageFont.truetype(font_path, 20)
                summary_font = ImageFont.truetype(font_path, 28)
            except Exception as e:
                logger.error(f"加载字体失败: {str(e)}")
                return jsonify({'error': '加载字体失败，请确保static目录下存在simsun.ttc字体文件'}), 500

            # 绘制表格边框
            def draw_cell(x, y, width, height, text, font, align='left', fill='black'):
                # 绘制边框
                draw.rectangle((x, y, x + width, y + height), outline='black', width=1)
                # 计算文本位置
                text_width = draw.textlength(text, font=font)
                if align == 'center':
                    text_x = x + (width - text_width) // 2
                elif align == 'right':
                    text_x = x + width - text_width - 10
                else:
                    text_x = x + 10
                text_y = y + (height - font.size) // 2
                # 绘制文本
                draw.text((text_x, text_y), text, fill=fill, font=font)
            
            # 绘制标题
            title = "发货单"
            title_width = draw.textlength(title, font=title_font)
            draw.text(((800 - title_width) // 2, 30), title, fill='black', font=title_font)
            
            # 绘制基本信息表格
            y = 100
            row_height = 40
            
            # 第一行
            draw_cell(50, y, 200, row_height, "发货物流", font, 'left')
            draw_cell(250, y, 200, row_height, delivery_order.logistics_company or "无", font, 'left')
            draw_cell(450, y, 150, row_height, "创建时间", font, 'left')
            draw_cell(600, y, 150, row_height, delivery_order.created_at.strftime('%Y/%m/%d'), font, 'left')
            
            # 第二行
            y += row_height
            draw_cell(50, y, 200, row_height, "客户名称", font, 'left')
            draw_cell(250, y, 500, row_height, delivery_order.customer_name, font, 'left')
            
            # 绘制商品表格标题
            y += row_height + 20
            draw.text((50, y), "商品明细", fill='black', font=font)
            y += 40
            
            # 表头
            header_height = 40
            # 新的列宽分配
            col_widths = {
                'name': 250,    # 商品名称加宽
                'color': 120,   # 颜色
                'quantity': 100, # 数量
                'price': 120,   # 单价加宽
                'total': 120    # 小计加宽
            }
            
            # 计算起始x坐标
            x = 50
            
            # 绘制表头
            draw_cell(x, y, col_widths['name'], header_height, "商品名称", font, 'center')
            x += col_widths['name']
            draw_cell(x, y, col_widths['color'], header_height, "颜色", font, 'center')
            x += col_widths['color']
            draw_cell(x, y, col_widths['quantity'], header_height, "数量", font, 'center')
            x += col_widths['quantity']
            draw_cell(x, y, col_widths['price'], header_height, "单价", font, 'center')
            x += col_widths['price']
            draw_cell(x, y, col_widths['total'], header_height, "小计", font, 'center')
            
            # 按商品ID和颜色分组汇总数量
            grouped_items = {}
            for item in delivery_items:
                key = (item.product_id, item.color)
                if key not in grouped_items:
                    grouped_items[key] = {
                        'product_id': item.product_id,
                        'color': item.color,
                        'quantity': 0
                    }
                grouped_items[key]['quantity'] += item.quantity

            # 绘制商品列表
            y += header_height
            total_amount = 0  # 重置总金额
            total_quantity = 0  # 重置总数量
            
            for item_info in grouped_items.values():
                product = Product.query.get(item_info['product_id'])
                if product:
                    # 从价格映射中获取单价信息
                    price_info = price_map.get((item_info['product_id'], item_info['color']), {
                        'price': 0,
                        'logo_price': 0,
                        'accessory_price': 0,
                        'packaging_price': 0
                    })
                    
                    # 计算总价（基础价格 + 加工费用）
                    unit_price = (price_info['price'] + price_info['logo_price'] + 
                                price_info['accessory_price'] + price_info['packaging_price'])
                    item_total = float(item_info['quantity'] * unit_price)
                    total_amount += item_total
                    total_quantity += item_info['quantity']
                    
                    row_height = 40
                    x = 50  # 重置x坐标到起始位置
                    
                    # 绘制每一列
                    draw_cell(x, y, col_widths['name'], row_height, product.name, small_font, 'left')
                    x += col_widths['name']
                    draw_cell(x, y, col_widths['color'], row_height, item_info['color'] or '', small_font, 'center')
                    x += col_widths['color']
                    draw_cell(x, y, col_widths['quantity'], row_height, str(item_info['quantity']), small_font, 'center')
                    x += col_widths['quantity']
                    draw_cell(x, y, col_widths['price'], row_height, f"¥{unit_price:.2f}", small_font, 'right')
                    x += col_widths['price']
                    draw_cell(x, y, col_widths['total'], row_height, f"¥{item_total:.2f}", small_font, 'right')
                    
                    y += row_height
            
            # 移除合计行，改为在右下角显示统计信息
            y += 40  # 添加一些间距
            
            
            # 绘制统计信息（右对齐）
            # 图片宽度为800，预留右边距20像素
            right_margin = 20
            right_x = 800 - right_margin
            
            # 绘制总包数（右对齐）
            text = f"总包数：{package_count or 1}包"
            text_width = draw.textlength(text, font=summary_font)
            draw.text((right_x - text_width, y), text, fill='black', font=summary_font)
            y += 50
            
            # 绘制商品总数（右对齐）
            text = f"商品总数：{total_quantity}件"
            text_width = draw.textlength(text, font=summary_font)
            draw.text((right_x - text_width, y), text, fill='black', font=summary_font)
            y += 50
            
            # 绘制货款总额（右对齐）
            text = f"货款总额：¥{int(total_amount)}"
            text_width = draw.textlength(text, font=summary_font)
            draw.text((right_x - text_width, y), text, fill='black', font=summary_font)
            y += 50
            
            # 绘制累计应付总额（右对齐）
            text = f"累计应付总额：¥{int(total_unpaid)}"
            text_width = draw.textlength(text, font=summary_font)
            draw.text((right_x - text_width, y), text, fill='black', font=summary_font)
            
            # 绘制备注
            if delivery_order.remark:
                y += 60  # 增加间距
                text = f"备注：{delivery_order.remark}"
                text_width = draw.textlength(text, font=title_font)
                draw.text((right_x - text_width, y), text, fill='black', font=title_font)
            
            # 将图片保存到内存中
            img_io = BytesIO()
            image.save(img_io, 'PNG')
            img_io.seek(0)
            
            # 获取云存储上传链接
            try:
                upload_url = 'http://api.weixin.qq.com/tcb/uploadfile'
                upload_params = {
                    'env': WX_ENV,
                    'path': f'delivery_images/delivery.png'
                }
                logger.info(f"请求云存储上传链接: {upload_params}")
                
                upload_response = requests.post(upload_url, json=upload_params)
                upload_data = upload_response.json()
                logger.info(f"获取上传链接响应: {upload_data}")
                
                if upload_data.get('errcode') != 0:
                    logger.error(f"获取上传链接失败: {upload_data}")
                    return jsonify({
                        'code': 500,
                        'message': '获取上传链接失败',
                        'data': upload_data
                    }), 500
                
            except Exception as e:
                logger.error(f"获取上传链接失败: {str(e)}")
                return jsonify({'error': f'获取上传链接失败: {str(e)}'}), 500
            
            # 上传文件到云存储
            try:
                cos_url = upload_data['url']
                files = {
                    'file': (f'delivery.png', img_io, 'image/png')
                }
                form_data = {
                    'key': f'delivery_images/delivery.png',
                    'Signature': upload_data['authorization'],
                    'x-cos-security-token': upload_data['token'],
                    'x-cos-meta-fileid': upload_data['file_id']
                }
                
                upload_result = requests.post(cos_url, data=form_data, files=files)
                
                logger.info(f"图片上传成功: file_id={upload_data['file_id']}")
                
                db.session.commit()
                
                # 返回图片访问地址
                return jsonify({
                    'code': 0,
                    'message': '发货单图片生成成功',
                    'data': {
                        'image_url': upload_data['file_id']
                    }
                })
                
            except Exception as e:
                logger.error(f"上传文件到云存储失败: {str(e)}")
                return jsonify({'error': f'上传文件到云存储失败: {str(e)}'}), 500
            
        except Exception as e:
            logger.error(f"图片生成过程发生错误: {str(e)}")
            return jsonify({'error': f'图片生成失败: {str(e)}'}), 500
            
    except Exception as e:
        error_msg = f"生成发货单图片失败: delivery_id={delivery_id}, error={str(e)}"
        logger.error(error_msg)
        return jsonify({'error': error_msg}), 500

@billing_bp.route('/delivery-orders/export', methods=['GET'])
@login_required
def export_delivery_orders(user_id):
    """导出发货单列表为Excel文件"""
    try:
        # 获取查询参数
        customer_id = request.args.get('customer_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 构建基础查询
        query = db.session.query(
            DeliveryOrder,
            User.nickname.label('customer_name'),
            db.func.sum(DeliveryItem.quantity).label('total_quantity'),
            db.func.sum(PurchaseOrderItem.price * DeliveryItem.quantity).label('total_amount'),
            DeliveryOrder.additional_fee.label('additional_fee')  # 添加附加费用字段
        ).join(
            DeliveryItem, DeliveryOrder.id == DeliveryItem.delivery_id
        ).join(
            PurchaseOrder, DeliveryItem.order_number == PurchaseOrder.order_number
        ).join(
            PurchaseOrderItem, 
            db.and_(
                PurchaseOrder.id == PurchaseOrderItem.order_id,
                PurchaseOrderItem.product_id == DeliveryItem.product_id,
                PurchaseOrderItem.color == DeliveryItem.color
            )
        ).join(
            User, DeliveryOrder.customer_id == User.id
        )
        
        # 添加筛选条件
        if customer_id:
            query = query.filter(DeliveryOrder.customer_id == customer_id).order_by(DeliveryOrder.created_at.desc())
        if start_date:
            try:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(DeliveryOrder.created_at >= start_datetime)
            except ValueError:
                return jsonify({'error': '开始日期格式错误，请使用YYYY-MM-DD格式'}), 400
        if end_date:
            try:
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                end_datetime = end_datetime + timedelta(days=1) - timedelta(seconds=1)
                query = query.filter(DeliveryOrder.created_at <= end_datetime)
            except ValueError:
                return jsonify({'error': '结束日期格式错误，请使用YYYY-MM-DD格式'}), 400
            
        # 分组并执行查询
        orders = query.group_by(DeliveryOrder.id, User.nickname).all()
        
        # 准备数据
        data = []
        for order, customer_name, total_quantity, total_amount, additional_fee in orders:
            # 状态文本映射
            status_text_map = {
                0: '已开单',
                1: '已发货',
                2: '已完成',
                3: '已取消',
                4: '异常'
            }
            
            # 计算实际总金额（商品金额 + 附加费用）
            actual_total = float(total_amount or 0) + float(additional_fee or 0)
            
            data.append({
                '发货单号': order.order_number,
                '客户名称': customer_name,
                '发货日期': order.created_at.strftime('%Y-%m-%d'),
                '商品总数': total_quantity or 0,
                '商品金额': float(total_amount or 0),
                '附加费用': float(additional_fee or 0),
                '货款总额': actual_total,
                '状态': status_text_map.get(order.status, '未知状态'),
                '物流公司': order.logistics_company or '',
                '物流单号': order.tracking_number or '',
                '备注': order.remark or ''
            })
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 设置列宽
        column_widths = {
            '发货单号': 20,
            '客户名称': 15,
            '发货日期': 12,
            '商品总数': 10,
            '商品金额': 12,
            '附加费用': 12,
            '货款总额': 12,
            '状态': 10,
            '物流公司': 15,
            '物流单号': 15,
            '备注': 30
        }
        
        # 创建Excel写入器
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='发货单列表')
            
            # 获取工作表
            worksheet = writer.sheets['发货单列表']
            
            # 设置列宽
            for col, width in column_widths.items():
                col_idx = df.columns.get_loc(col)
                worksheet.column_dimensions[chr(65 + col_idx)].width = width
            
            # 设置表头样式
            for col in range(len(df.columns)):
                cell = worksheet.cell(row=1, column=col+1)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # 设置数据样式
            for row in range(2, len(df) + 2):
                for col in range(len(df.columns)):
                    cell = worksheet.cell(row=row, column=col+1)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # 设置金额列的格式
                    if df.columns[col] in ['商品金额', '附加费用', '货款总额']:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.number_format = '¥#,##0.00'
        
        output.seek(0)
        
        # 生成文件名
        filename = f'发货单列表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f'导出发货单列表失败: {str(e)}')
        print(f'错误追踪:\n{traceback.format_exc()}')
        return jsonify({'error': '导出发货单列表失败'}), 500

@billing_bp.route('/user-product-prices/import', methods=['POST'])
@login_required
def import_user_product_prices(user_id):
    """导入用户商品价格"""
    try:
        print("[DEBUG] 开始导入用户商品价格")
        # 检查是否有文件上传
        if 'file' not in request.files:
            print("[ERROR] 没有上传文件")
            return jsonify({'error': '没有上传文件'}), 400
            
        file = request.files['file']
        print(f"[DEBUG] 收到文件: {file.filename}")
        if not file.filename.endswith('.xlsx'):
            print("[ERROR] 文件格式不支持")
            return jsonify({'error': '只支持.xlsx格式的Excel文件'}), 400
            
        # 读取Excel文件
        print("[DEBUG] 开始读取Excel文件")
        df = pd.read_excel(file)
        print(f"[DEBUG] Excel文件读取完成，共{len(df)}行数据")
        
        # 检查必要的列是否存在
        required_columns = ['客户名称', '商品', '价格']
        if not all(col in df.columns for col in required_columns):
            print(f"[ERROR] Excel文件缺少必要列，需要: {required_columns}")
            return jsonify({'error': f'Excel文件必须包含以下列：{", ".join(required_columns)}'}), 400
            
        # 验证数据类型
        try:
            print("[DEBUG] 开始验证价格列数据类型")
            df['价格'] = df['价格'].astype(float)
            print("[DEBUG] 价格列数据类型验证通过")
        except Exception as e:
            print(f"[ERROR] 价格列数据类型验证失败: {str(e)}")
            return jsonify({'error': f'价格列必须为数字：{str(e)}'}), 400
            
        # 获取所有客户名称和商品名称
        customer_names = df['客户名称'].unique()
        product_names = df['商品'].unique()
        print(f"[DEBUG] 共发现{len(customer_names)}个客户，{len(product_names)}个商品")
        
        # 查询数据库中存在的客户和商品
        print("[DEBUG] 开始查询数据库中的客户和商品")
        existing_customers = User.query.filter(
            User.nickname.in_(customer_names)
        ).all()
        existing_products = Product.query.filter(Product.name.in_(product_names)).all()
        print(f"[DEBUG] 数据库中存在{len(existing_customers)}个客户，{len(existing_products)}个商品")
        
        # 创建名称到ID的映射
        customer_name_to_id = {customer.nickname: customer.id for customer in existing_customers}
        product_name_to_id = {product.name: product.id for product in existing_products}
        
        # 检查不存在的客户和商品
        invalid_customers = set(customer_names) - set(customer_name_to_id.keys())
        invalid_products = set(product_names) - set(product_name_to_id.keys())
        
        if invalid_customers:
            print(f"[ERROR] 发现不存在的客户: {invalid_customers}")
            return jsonify({'error': f'以下客户不存在：{", ".join(invalid_customers)}'}), 400
            
        if invalid_products:
            print(f"[WARNING] 发现不存在的商品: {invalid_products}，将跳过这些商品")
            
        # 开始导入数据
        print("[DEBUG] 开始导入数据")
        success_count = 0
        error_count = 0
        skip_count = 0
        error_messages = []
        skip_messages = []
        
        for _, row in df.iterrows():
            try:
                customer_id = customer_name_to_id[row['客户名称']]
                product_name = row['商品']
                
                # 检查商品是否存在
                if product_name not in product_name_to_id:
                    skip_count += 1
                    skip_msg = f'第{_+2}行跳过：商品"{product_name}"不存在'
                    print(f"[WARNING] {skip_msg}")
                    skip_messages.append(skip_msg)
                    continue
                    
                product_id = product_name_to_id[product_name]
                
                # 检查是否已存在相同的记录
                existing_price = UserProductPrice.query.filter_by(
                    user_id=customer_id,
                    product_id=product_id
                ).first()
                
                if existing_price:
                    print(f"[DEBUG] 更新现有价格记录: 客户ID={customer_id}, 商品ID={product_id}")
                    # 更新现有记录
                    existing_price.custom_price = row['价格']
                else:
                    print(f"[DEBUG] 创建新价格记录: 客户ID={customer_id}, 商品ID={product_id}")
                    # 创建新记录
                    new_price = UserProductPrice(
                        user_id=customer_id,
                        product_id=product_id,
                        custom_price=row['价格']
                    )
                    db.session.add(new_price)
                
                success_count += 1
                
            except Exception as e:
                error_count += 1
                error_msg = f'第{_+2}行导入失败：{str(e)}'
                print(f"[ERROR] {error_msg}")
                error_messages.append(error_msg)
                
        # 提交事务
        print(f"[DEBUG] 开始提交事务，成功导入{success_count}条，失败{error_count}条，跳过{skip_count}条")
        db.session.commit()
        print("[DEBUG] 事务提交成功")
        
        return jsonify({
            'message': '导入完成',
            'success_count': success_count,
            'error_count': error_count,
            'skip_count': skip_count,
            'errors': error_messages,
            'skips': skip_messages
        })
        
    except Exception as e:
        print(f"[ERROR] 导入用户商品价格失败: {str(e)}")
        db.session.rollback()
        return jsonify({'error': f'导入失败：{str(e)}'}), 500

@billing_bp.route('/purchase_orders/update_prices', methods=['POST'])
@login_required
def update_purchase_order_prices(user_id):
    """更新指定客户在指定时间段的采购单价格"""
    try:
        print("[DEBUG] 开始更新采购单价格")
        data = request.get_json()
        
        # 获取请求参数
        customer_id = data.get('customer_id')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        if not all([customer_id, start_date, end_date]):
            print("[ERROR] 缺少必要参数")
            return jsonify({'error': '缺少必要参数：customer_id, start_date, end_date'}), 400
            
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
        except ValueError as e:
            print(f"[ERROR] 日期格式错误: {str(e)}")
            return jsonify({'error': '日期格式错误，请使用YYYY-MM-DD格式'}), 400
            
        print(f"[DEBUG] 查询参数: 客户ID={customer_id}, 开始日期={start_date}, 结束日期={end_date}")
        
        # 获取客户的所有采购单
        purchase_orders = PurchaseOrder.query.filter(
            PurchaseOrder.user_id == customer_id,
            PurchaseOrder.created_at.between(start_date, end_date)
        ).all()
        
        if not purchase_orders:
            print("[WARNING] 未找到符合条件的采购单")
            return jsonify({'message': '未找到符合条件的采购单'}), 200
            
        print(f"[DEBUG] 找到{len(purchase_orders)}个采购单")
        
        # 获取客户的所有商品价格
        user_prices = UserProductPrice.query.filter_by(user_id=customer_id).all()
        price_map = {price.product_id: price.custom_price for price in user_prices}
        
        if not price_map:
            print("[WARNING] 未找到客户商品价格记录")
            return jsonify({'message': '未找到客户商品价格记录'}), 200
            
        print(f"[DEBUG] 找到{len(price_map)}个商品价格记录")
        
        # 更新采购单价格
        updated_count = 0
        error_count = 0
        error_messages = []
        
        for order in purchase_orders:
            try:
                # 获取采购单的所有商品
                order_items = PurchaseOrderItem.query.filter_by(order_id=order.id).all()
                
                for item in order_items:
                    # 如果指定了商品名称，则只更新匹配的商品
                    if data.get('product_name'):
                        product = Product.query.get(item.product_id)
                        if not product or product.name != data.get('product_name'):
                            continue
                            
                    if item.product_id in price_map:
                        # 更新商品价格
                        old_price = item.price
                        new_price = price_map[item.product_id]
                        item.price = new_price
                        print(f"[DEBUG] 更新价格: 订单ID={order.id}, 商品ID={item.product_id}, 原价={old_price}, 新价={new_price}")
                        updated_count += 1
                
                # 重新计算采购单总金额
                order.total_amount = sum(item.price * item.quantity for item in order_items)
                
            except Exception as e:
                error_count += 1
                error_msg = f'更新采购单{order.id}失败: {str(e)}'
                print(f"[ERROR] {error_msg}")
                error_messages.append(error_msg)
                continue
        
        # 提交事务
        print(f"[DEBUG] 开始提交事务，成功更新{updated_count}个商品价格，失败{error_count}个")
        db.session.commit()
        print("[DEBUG] 事务提交成功")
        
        return jsonify({
            'message': '更新完成',
            'updated_count': updated_count,
            'error_count': error_count,
            'errors': error_messages
        })
        
    except Exception as e:
        print(f"[ERROR] 更新采购单价格失败: {str(e)}")
        db.session.rollback()
        return jsonify({'error': f'更新失败：{str(e)}'}), 500

@billing_bp.route('/orders/delete_all', methods=['POST'])
@login_required
def delete_all_orders(user_id):
    """删除指定客户的所有订单（采购单和发货单）"""
    try:
        print("[DEBUG] 开始删除客户所有订单")
        data = request.get_json()
        customer_name = data.get('customer_name')
        
        if not customer_name:
            print("[ERROR] 缺少客户名称参数")
            return jsonify({'error': '缺少客户名称参数'}), 400
            
        # 查询客户信息
        customer = User.query.filter_by(nickname=customer_name).first()
        if not customer:
            print(f"[ERROR] 未找到客户: {customer_name}")
            return jsonify({'error': f'未找到客户: {customer_name}'}), 404
            
        # 开始事务
        db.session.begin()
        
        # 1. 删除发货单相关数据
        # 先查询所有发货单ID
        delivery_orders = DeliveryOrder.query.filter_by(customer_id=customer.id).all()
        delivery_ids = [order.id for order in delivery_orders]
        
        if delivery_ids:
            # 删除发货单商品
            DeliveryItem.query.filter(DeliveryItem.delivery_id.in_(delivery_ids)).delete()
            print(f"[DEBUG] 已删除发货单商品记录: {len(delivery_ids)}个发货单")
            
            # 删除发货单
            DeliveryOrder.query.filter(DeliveryOrder.id.in_(delivery_ids)).delete()
            print(f"[DEBUG] 已删除发货单记录: {len(delivery_ids)}个")
        
        # 2. 删除采购单相关数据
        # 先查询所有采购单ID
        purchase_orders = PurchaseOrder.query.filter_by(user_id=customer.id).all()
        purchase_ids = [order.id for order in purchase_orders]
        
        if purchase_ids:
            # 删除采购单商品
            PurchaseOrderItem.query.filter(PurchaseOrderItem.order_id.in_(purchase_ids)).delete()
            print(f"[DEBUG] 已删除采购单商品记录: {len(purchase_ids)}个采购单")
            
            # 删除采购单
            PurchaseOrder.query.filter(PurchaseOrder.id.in_(purchase_ids)).delete()
            print(f"[DEBUG] 已删除采购单记录: {len(purchase_ids)}个")
        
        # 3. 删除收款记录
        Payment.query.filter_by(customer_id=customer.id).delete()
        print(f"[DEBUG] 已删除收款记录")
        
        # 提交事务
        db.session.commit()
        
        return jsonify({
            'message': '删除成功',
            'deleted_delivery_orders': len(delivery_ids),
            'deleted_purchase_orders': len(purchase_ids)
        })
        
    except Exception as e:
        print(f"[ERROR] 删除订单失败: {str(e)}")
        print(f"错误追踪:\n{traceback.format_exc()}")
        db.session.rollback()
        return jsonify({'error': f'删除订单失败: {str(e)}'}), 500

@billing_bp.route('/user-product-prices', methods=['GET'])
@login_required
def get_user_product_prices(user_id):
    """获取用户商品价格列表"""
    try:
        # 获取查询参数
        customer_id = request.args.get('customer_id')
        product_id = request.args.get('product_id')
        keyword = request.args.get('keyword')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))

        # 构建基础查询
        query = db.session.query(
            UserProductPrice,
            User.nickname.label('customer_name'),
            Product.name.label('product_name')
        ).join(
            User, UserProductPrice.user_id == User.id
        ).join(
            Product, UserProductPrice.product_id == Product.id
        )

        # 添加筛选条件
        if customer_id:
            query = query.filter(UserProductPrice.user_id == customer_id)
        if product_id:
            query = query.filter(UserProductPrice.product_id == product_id)
        if keyword:
            query = query.filter(
                db.or_(
                    User.nickname.ilike(f'%{keyword}%'),
                    Product.name.ilike(f'%{keyword}%')
                )
            )

        # 获取总数
        total = query.count()

        # 分页
        query = query.offset((page - 1) * page_size).limit(page_size)

        # 执行查询
        results = query.all()

        # 构建返回数据
        items = []
        for price, customer_name, product_name in results:
            items.append({
                'id': price.id,
                'customer_id': price.user_id,
                'customer_name': customer_name,
                'product_id': price.product_id,
                'product_name': product_name,
                'price': float(price.custom_price) if price.custom_price else None,
                'created_at': price.created_at.strftime('%Y-%m-%d %H:%M:%S') if price.created_at else None,
                'updated_at': price.updated_at.strftime('%Y-%m-%d %H:%M:%S') if price.updated_at else None
            })

        return jsonify({
            'code': 0,
            'data': {
                'items': items,
                'total': total,
                'page': page,
                'page_size': page_size
            }
        })

    except Exception as e:
        print(f"[ERROR] 获取用户商品价格列表失败: {str(e)}")
        return jsonify({
            'code': -1,
            'message': f'获取用户商品价格列表失败: {str(e)}'
        })

@billing_bp.route('/user-product-prices', methods=['POST'])
@login_required
def submit_user_product_price(user_id):
    """提交用户商品价格"""
    try:
        data = request.get_json()
        
        # 验证必要参数
        required_fields = ['user_id', 'product_id', 'price']
        if not all(field in data for field in required_fields):
            return jsonify({
                'code': -1,
                'message': '缺少必要参数：user_id, product_id, price'
            })
            
        # 验证价格是否为有效数字
        try:
            price = float(data['price'])
            if price < 0:
                raise ValueError('价格不能为负数')
        except ValueError as e:
            return jsonify({
                'code': -1,
                'message': f'价格格式错误：{str(e)}'
            })
            
        # 检查用户和商品是否存在
        user = User.query.get(data['user_id'])
        if not user:
            return jsonify({
                'code': -1,
                'message': '用户不存在'
            })
            
        product = Product.query.get(data['product_id'])
        if not product:
            return jsonify({
                'code': -1,
                'message': '商品不存在'
            })
            
        # 检查是否已存在价格记录
        existing_price = UserProductPrice.query.filter_by(
            user_id=data['user_id'],
            product_id=data['product_id']
        ).first()
        
        if existing_price:
            # 更新现有价格
            existing_price.custom_price = price
            existing_price.updated_at = datetime.now()
            message = '价格更新成功'
        else:
            # 创建新价格记录
            new_price = UserProductPrice(
                user_id=data['user_id'],
                product_id=data['product_id'],
                custom_price=price,
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            db.session.add(new_price)
            message = '价格创建成功'
            
        # 提交事务
        db.session.commit()
        
        return jsonify({
            'code': 0,
            'message': message
        })
        
    except Exception as e:
        print(f"[ERROR] 提交用户商品价格失败: {str(e)}")
        db.session.rollback()
        return jsonify({
            'code': -1,
            'message': f'提交用户商品价格失败: {str(e)}'
        })

@billing_bp.route('/user-product-prices/all', methods=['GET'])
@login_required
def get_all_user_product_prices(user_id):
    """获取指定客户的所有商品价格"""
    try:
        # 获取查询参数
        customer_id = request.args.get('customer_id')
        
        if not customer_id:
            return jsonify({
                'code': -1,
                'message': '缺少必要参数：customer_id'
            })
            
        # 构建查询
        prices = db.session.query(
            UserProductPrice,
            Product.name.label('product_name'),
            Product.id.label('product_id')
        ).join(
            Product, UserProductPrice.product_id == Product.id
        ).filter(
            UserProductPrice.user_id == customer_id
        ).all()
        
        # 构建返回数据
        result = []
        for price, product_name, product_id in prices:
            result.append({
                'productName': product_name,
                'productId': product_id,
                'price': float(price.custom_price) if price.custom_price else 0
            })
            
        return jsonify({
            'code': 0,
            'prices': result
        })
        
    except Exception as e:
        print(f"[ERROR] 获取客户商品价格失败: {str(e)}")
        return jsonify({
            'code': -1,
            'message': f'获取客户商品价格失败: {str(e)}'
        })

